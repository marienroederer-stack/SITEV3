"""Import d'un fichier de listing d'appels (Excel ou CSV) dans la base locale.

Règles de filtrage (fixées avec le client) :
 - uniquement les appels entrants aboutis (Type = entrant, Raison du rejet = normal ou vide)
 - la durée retenue est la colonne "Comm" (temps de communication réel)
 - les appels de moins de 8 secondes de Comm ne sont pas comptabilisés
 - dédoublonnage entre imports successifs via "Call Id"
"""

import csv
import hashlib
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from sqlite3 import Connection
from typing import Optional

import openpyxl

from . import db

# Colonnes recherchées dans le fichier source, avec leurs libellés possibles
# (les exports peuvent légèrement varier). Comparaison faite sans accents / casse.
COLUMN_CANDIDATES = {
    "date": ["date"],
    "heure": ["heure"],
    "type": ["type"],
    "numero_appelant": ["numero appelant"],
    "numero_appele": ["numero appele"],
    "nom_appele": ["nom appele"],
    "comm": ["comm"],
    "raison_rejet": ["raison du rejet"],
    "call_id": ["call id", "callid"],
    "tag": ["tag"],
}


def _normalize(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    return text.strip().lower()


def _build_header_index(headers: list, column_candidates: Optional[dict] = None) -> dict:
    normalized = {_normalize(h): i for i, h in enumerate(headers) if h is not None}
    index = {}
    for field_name, candidates in (column_candidates or COLUMN_CANDIDATES).items():
        for candidate in candidates:
            if candidate in normalized:
                index[field_name] = normalized[candidate]
                break
    return index


def parse_duration_seconds(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, timedelta):
        return int(value.total_seconds())
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return 0
    if ":" in text:
        parts = [int(p) for p in text.split(":")]
        while len(parts) < 3:
            parts.insert(0, 0)
        h, m, s = parts[-3:]
        return h * 3600 + m * 60 + s
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_datetime(value, date_fallback=None) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    text = str(value).strip() if value is not None else ""
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    # Cas où "Heure" ne contient qu'une heure sans date : on complète avec la date de la colonne "Date".
    if date_fallback is not None:
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = datetime.strptime(text, fmt).time()
                d = parse_datetime(date_fallback)
                if d is not None:
                    return datetime.combine(d.date(), t)
            except ValueError:
                continue
    return None


@dataclass
class ImportResult:
    total_rows: int = 0
    inserted: int = 0
    duplicates: int = 0
    filtered_out: int = 0
    invalid_rows: int = 0
    new_clients: list = field(default_factory=list)


def _iter_rows_xlsx(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    for row in rows:
        yield headers, row


def _iter_rows_csv(path: Path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(f, dialect)
        headers = next(reader)
        for row in reader:
            yield headers, row


def import_file(conn: Connection, path: Path) -> ImportResult:
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        row_iter = _iter_rows_xlsx(path)
    elif path.suffix.lower() == ".csv":
        row_iter = _iter_rows_csv(path)
    else:
        raise ValueError(f"Format de fichier non supporté : {path.suffix}")

    result = ImportResult()
    idx = None
    known_clients = {c.numero_appele for c in db.list_clients(conn)}

    for headers, row in row_iter:
        if idx is None:
            idx = _build_header_index(headers)
            missing = [f for f in ("heure", "type", "numero_appele", "nom_appele", "comm") if f not in idx]
            if missing:
                raise ValueError(
                    "Colonnes manquantes dans le fichier : " + ", ".join(missing)
                )

        result.total_rows += 1

        def cell(field_name):
            i = idx.get(field_name)
            return row[i] if i is not None and i < len(row) else None

        type_val = _normalize(cell("type") or "")
        raison_val = _normalize(cell("raison_rejet") or "")
        if type_val != "entrant" or raison_val not in ("", "normal"):
            result.filtered_out += 1
            continue

        comm_seconds = parse_duration_seconds(cell("comm"))
        if comm_seconds < 8:
            result.filtered_out += 1
            continue

        numero_appele = str(cell("numero_appele") or "").strip()
        nom_appele = str(cell("nom_appele") or "").strip()
        if not numero_appele or not nom_appele:
            result.invalid_rows += 1
            continue

        call_dt = parse_datetime(cell("heure"), date_fallback=cell("date"))
        if call_dt is None:
            result.invalid_rows += 1
            continue

        call_id = str(cell("call_id") or "").strip()
        if not call_id:
            # Pas d'identifiant fourni : on en fabrique un stable pour permettre le dédoublonnage.
            raw = f"{numero_appele}|{call_dt.isoformat()}|{cell('numero_appelant')}"
            call_id = hashlib.sha1(raw.encode("utf-8")).hexdigest()

        client = db.get_or_create_client(conn, numero_appele, nom_appele)
        if numero_appele not in known_clients:
            known_clients.add(numero_appele)
            result.new_clients.append(client)

        inserted = db.insert_call(
            conn,
            call_id=call_id,
            numero_appele=numero_appele,
            numero_appelant=str(cell("numero_appelant") or "").strip(),
            date_heure=call_dt,
            comm_seconds=comm_seconds,
            tag=str(cell("tag") or "").strip(),
        )
        if inserted:
            result.inserted += 1
        else:
            result.duplicates += 1

    conn.commit()
    return result
