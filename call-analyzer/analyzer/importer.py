"""Import d'un fichier de listing d'appels (Excel ou CSV) dans la base locale.

Règle de filtrage : seuls les appels entrants (Type = entrant) sont retenus — cet
outil est une analyse interne globale (volume, TAG, attente...), pas un rapport de
facturation, donc aucun filtre de durée minimale ni de raison de rejet n'est
appliqué : tous les appels entrants comptent, y compris les appels courts ou non
aboutis.

Dédoublonnage entre imports successifs via la colonne "Call Id" (un même fichier ou
une période qui se chevauche peut être réimporté sans créer de doublons).

Les nouveaux SDA et les nouveaux logins opérateur rencontrés dans un fichier sont
ajoutés automatiquement aux répertoires (Listing clients / Listing opérateurs). Si
le nom ou le code affaire d'un SDA déjà connu change dans un nouvel import, la
fiche est mise à jour mais l'historique des appels reste attaché au même SDA.
"""

import csv
import hashlib
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from sqlite3 import Connection
from typing import Optional

import openpyxl

from . import db

COLUMN_CANDIDATES = {
    "date": ["date"],
    "heure": ["heure"],
    "type": ["type"],
    "numero_appelant": ["numero appelant"],
    "numero_appele": ["numero appele"],
    "identifiant_appele": ["identifiant appele"],
    "nom_appele": ["nom appele"],
    "duree_totale": ["duree totale"],
    "annonce": ["annonce"],
    "file": ["file"],
    "sonnerie": ["sonnerie"],
    "comm": ["comm"],
    "tag": ["tag"],
    "raison_rejet": ["raison du rejet"],
    "call_id": ["call id", "callid"],
    "priorite": ["priorite"],
    "code_affaire": ["code affaire"],
}

REQUIRED_FIELDS = ("heure", "type", "numero_appele", "nom_appele")


def _normalize(text) -> str:
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    return text.strip().lower()


def _build_header_index(headers: list) -> dict:
    normalized = {_normalize(h): i for i, h in enumerate(headers) if h is not None}
    index = {}
    for field_name, candidates in COLUMN_CANDIDATES.items():
        for candidate in candidates:
            if candidate in normalized:
                index[field_name] = normalized[candidate]
                break
    return index


def parse_duration_seconds(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, timedelta):
        return int(value.total_seconds())
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return 0
    if ":" in text:
        parts = [int(p) for p in text.split(":")]
        while len(parts) < 3:
            parts.insert(0, 0)
        h, m, s = parts[-3:]
        return h * 3600 + m * 60 + s
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_datetime(value, date_fallback=None) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    text = str(value).strip() if value is not None else ""
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    # Cas où "Heure" ne contient qu'une heure sans date : on complète avec la colonne "Date".
    if date_fallback is not None:
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = datetime.strptime(text, fmt).time()
                d = parse_datetime(date_fallback)
                if d is not None:
                    return datetime.combine(d.date(), t)
            except ValueError:
                continue
    return None


@dataclass
class ImportResult:
    total_rows: int = 0
    inserted: int = 0
    duplicates: int = 0
    filtered_out: int = 0
    invalid_rows: int = 0
    new_clients: list = field(default_factory=list)
    new_operators: list = field(default_factory=list)


def _iter_rows_xlsx(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    for row in rows:
        yield headers, row


def _iter_rows_csv(path: Path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(f, dialect)
        headers = next(reader)
        for row in reader:
            yield headers, row


def import_file(conn: Connection, path: Path) -> ImportResult:
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        row_iter = _iter_rows_xlsx(path)
    elif path.suffix.lower() == ".csv":
        row_iter = _iter_rows_csv(path)
    else:
        raise ValueError(f"Format de fichier non supporté : {path.suffix}")

    result = ImportResult()
    idx = None
    known_clients = {c.sda for c in db.list_clients(conn)}
    known_operators = {o.login for o in db.list_operators(conn)}
    import_id = None

    for headers, row in row_iter:
        if idx is None:
            idx = _build_header_index(headers)
            missing = [f for f in REQUIRED_FIELDS if f not in idx]
            if missing:
                raise ValueError("Colonnes manquantes dans le fichier : " + ", ".join(missing))
            import_id = db.record_import(conn, path.name, result)

        result.total_rows += 1

        def cell(field_name):
            i = idx.get(field_name)
            return row[i] if i is not None and i < len(row) else None

        type_val = _normalize(cell("type") or "")
        if type_val != "entrant":
            result.filtered_out += 1
            continue

        numero_appele = str(cell("numero_appele") or "").strip()
        nom_appele = str(cell("nom_appele") or "").strip()
        if not numero_appele:
            result.invalid_rows += 1
            continue

        call_dt = parse_datetime(cell("heure"), date_fallback=cell("date"))
        if call_dt is None:
            result.invalid_rows += 1
            continue

        code_affaire = str(cell("code_affaire") or "").strip()
        client, created = db.get_or_create_client(conn, numero_appele, nom_appele, code_affaire)
        if created:
            known_clients.add(numero_appele)
            result.new_clients.append(client)

        operateur = str(cell("identifiant_appele") or "").strip()
        if operateur and operateur not in known_operators:
            _, op_created = db.get_or_create_operator(conn, operateur)
            if op_created:
                known_operators.add(operateur)
                result.new_operators.append(operateur)

        call_id = str(cell("call_id") or "").strip()
        if not call_id:
            raw = f"{numero_appele}|{call_dt.isoformat()}|{cell('numero_appelant')}"
            call_id = hashlib.sha1(raw.encode("utf-8")).hexdigest()

        inserted = db.insert_call(
            conn,
            call_id=call_id,
            date_heure=call_dt.isoformat(),
            sda=numero_appele,
            operateur=operateur,
            numero_appelant=str(cell("numero_appelant") or "").strip(),
            code_affaire_row=code_affaire,
            tag=str(cell("tag") or "").strip(),
            priorite=str(cell("priorite") or "").strip(),
            raison_rejet=str(cell("raison_rejet") or "").strip(),
            duree_totale_seconds=parse_duration_seconds(cell("duree_totale")),
            annonce_seconds=parse_duration_seconds(cell("annonce")),
            file_seconds=parse_duration_seconds(cell("file")),
            sonnerie_seconds=parse_duration_seconds(cell("sonnerie")),
            comm_seconds=parse_duration_seconds(cell("comm")),
            import_id=import_id,
        )
        if inserted:
            result.inserted += 1
        else:
            result.duplicates += 1

    if import_id is not None:
        conn.execute(
            "UPDATE imports SET rows_total = ?, rows_inserted = ?, rows_duplicates = ?, "
            "rows_filtered = ?, rows_invalid = ?, new_clients = ?, new_operators = ? WHERE id = ?",
            (
                result.total_rows,
                result.inserted,
                result.duplicates,
                result.filtered_out,
                result.invalid_rows,
                len(result.new_clients),
                len(result.new_operators),
                import_id,
            ),
        )
    conn.commit()
    return result
